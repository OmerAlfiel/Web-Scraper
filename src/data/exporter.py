"""
Data Exporter Module

This module handles the exporting of cleaned data to various formats,
primarily Excel files.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import EXCEL_CONFIG
import logging
from datetime import datetime


def save_to_excel(data, file_path, sheet_name=None):
    """
    Saves the cleaned data to an Excel file with formatting.

    Parameters:
    data (DataFrame or list of dict): The cleaned data to be saved.
    file_path (str): The path where the Excel file will be saved.
    sheet_name (str, optional): The name of the worksheet.
    """
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
    
    try:
        # Convert to DataFrame if it's a list of dictionaries
        if not isinstance(data, pd.DataFrame):
            data = pd.DataFrame(data)
        
        if data.empty:
            logging.warning("No data to export to Excel")
            # Create an empty file with headers
            workbook = Workbook()
            worksheet = workbook.active
            if sheet_name:
                worksheet.title = sheet_name
            
            # Add headers
            headers = ["Project Name", "Project Location", "Project Type", 
                       "Contact Name", "Mobile Number", "Source URL"]
            worksheet.append(headers)
            apply_header_style(worksheet, len(headers))
            
            workbook.save(file_path)
            logging.info(f"Empty Excel file saved to: {file_path}")
            return
        
        # Create a new Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        if sheet_name:
            worksheet.title = sheet_name
        
        # Convert DataFrame to rows and add to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True)):
            worksheet.append(row)
            
            # Apply styles
            if r_idx == 0:
                apply_header_style(worksheet, len(row))
            else:
                status = row[data.columns.get_loc('Extraction Status')] if 'Extraction Status' in data.columns else None
                apply_row_style(worksheet, r_idx + 1, len(row), r_idx % 2 == 0, status)
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 12)  # Min width of 12
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # Max width of 50
        
        # Add metadata in a separate worksheet
        add_metadata_worksheet(workbook, data)
        
        # Save the workbook
        workbook.save(file_path)
        logging.info(f"Excel file saved to: {file_path}")
    
    except Exception as e:
        logging.error(f"Error saving Excel file: {str(e)}")
        raise


def apply_header_style(worksheet, num_columns):
    """Apply styling to header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, num_columns + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        

def apply_row_style(worksheet, row_idx, num_columns, is_alternate_row, status=None):
    """Apply styling to data rows."""
    cell_value = worksheet.cell(row=row_idx, column=1).value
    
    # Check if this is a timestamp header row
    if cell_value and isinstance(cell_value, str) and "--- Data Collected on" in cell_value:
        # Style for timestamp header - dark blue with white text
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12)
        alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        # Apply styles to the full row
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border
            
        # Merge the cells for the timestamp row if there are multiple columns
        if num_columns > 1:
            worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=num_columns)
        
        return

    # Normal row styling as before
    if status and "Failed" in status:
        # Light red for failed extractions
        fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    elif status and "Metadata" in status:
        # Light yellow for partial metadata usage
        fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    elif is_alternate_row:
        fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    else:
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Define borders
    border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Define alignment
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Apply styles to each cell in the row
    for col in range(1, num_columns + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        cell.fill = fill
        cell.border = border
        cell.alignment = alignment


def add_metadata_worksheet(workbook, data):
    """Add a metadata worksheet with summary information."""
    metadata_sheet = workbook.create_sheet(title="Metadata")
    
    # Add timestamp
    metadata_sheet.append(["Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    # Add record count
    metadata_sheet.append(["Total Records:", len(data)])
    
    # Add extraction status summary
    if 'Extraction Status' in data.columns:
        metadata_sheet.append([])
        metadata_sheet.append(["Extraction Status Summary:"])
        status_counts = data['Extraction Status'].value_counts().reset_index()
        status_counts.columns = ['Extraction Status', 'Count']
        
        metadata_sheet.append(["Extraction Status", "Count"])
        for _, row in status_counts.iterrows():
            metadata_sheet.append([row['Extraction Status'], row['Count']])
    
    # Add legend
    metadata_sheet.append([])
    metadata_sheet.append(["Color Legend:"])
    metadata_sheet.append(["White/Blue", "Successfully extracted data"])
    metadata_sheet.append(["Yellow", "Partial data from metadata"])
    metadata_sheet.append(["Red", "Failed extraction - using metadata only"])
    
    # Adjust column widths
    for column in metadata_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(max_length + 2, 15)
        metadata_sheet.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(data):
    """
    Export cleaned data to Excel file using configuration settings.
    Appends new data with timestamps rather than overwriting.
    """
    try:
        # Get configuration
        output_file = EXCEL_CONFIG.get('output_file', 'scraped_data.xlsx')
        sheet_name = EXCEL_CONFIG.get('sheet_name', 'Projects')
        
        # Check if output directory is specified
        output_dir = EXCEL_CONFIG.get('output_dir', '')
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            output_file = os.path.join(output_dir, output_file)
        
        # Convert to DataFrame if needed
        if not isinstance(data, pd.DataFrame):
            data = pd.DataFrame(data)
            
        # Check if file exists
        if os.path.exists(output_file):
            # Load existing workbook
            try:
                existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
                
                # Add timestamp header row before new data
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                header_df = pd.DataFrame([{"Project Name": f"--- Data Collected on {timestamp} ---"}])
                
                # Set all other columns to empty in the header row
                for col in data.columns:
                    if col != "Project Name" and col in header_df.columns:
                        header_df[col] = ""
                
                # Combine: existing data + timestamp header + new data
                combined_df = pd.concat([existing_df, header_df, data], ignore_index=True)
                
                # Save the combined data
                save_to_excel(combined_df, output_file, sheet_name)
                
            except Exception as e:
                logging.error(f"Error reading existing Excel file: {str(e)}")
                logging.warning("Creating new Excel file instead")
                save_to_excel(data, output_file, sheet_name)
        else:
            # File doesn't exist, create new one with timestamp header
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            header_df = pd.DataFrame([{"Project Name": f"--- Data Collected on {timestamp} ---"}])
            
            # Set all other columns to empty in the header row
            for col in data.columns:
                if col != "Project Name" and col in header_df.columns:
                    header_df[col] = ""
            
            # Combine header with data
            combined_df = pd.concat([header_df, data], ignore_index=True)
            save_to_excel(combined_df, output_file, sheet_name)
        
        return output_file
        
    except Exception as e:
        logging.error(f"Error in export_to_excel: {str(e)}")
        # Fallback to a simple filename in case of error
        fallback_file = 'scraped_data_fallback.xlsx'
        try:
            # Simplified save without fancy formatting
            if isinstance(data, pd.DataFrame):
                data.to_excel(fallback_file, index=False)
            else:
                pd.DataFrame(data).to_excel(fallback_file, index=False)
            logging.warning(f"Used fallback Excel export to: {fallback_file}")
            return fallback_file
        except:
            logging.error("Failed to save Excel file even with fallback method")
            return None


def export_to_csv(data, file_path=None):
    """
    Export data to CSV file.
    
    Parameters:
    data (list of dict or DataFrame): The data to be exported.
    file_path (str, optional): The path where the CSV file will be saved.
                              If not provided, uses the default location.
    
    Returns:
    str: The path to the saved CSV file.
    """
    try:
        if file_path is None:
            file_path = 'data/scraped_data.csv'
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # Convert to DataFrame if it's a list of dictionaries
        if not isinstance(data, pd.DataFrame):
            data = pd.DataFrame(data)
        
        # Save to CSV
        data.to_csv(file_path, index=False, encoding='utf-8')
        logging.info(f"CSV file saved to: {file_path}")
        
        return file_path
    
    except Exception as e:
        logging.error(f"Error exporting to CSV: {str(e)}")
        return None


if __name__ == "__main__":
    # Test with sample data
    sample_data = [
        {
            "Project Name": "Test Project 1",
            "Project Location": "New York",
            "Project Type": "Web Development",
            "Contact Name": "John Doe",
            "Mobile Number": "123-456-7890",
            "Source URL": "https://example.com/1"
        },
        {
            "Project Name": "Test Project 2",
            "Project Location": "San Francisco",
            "Project Type": "Mobile App",
            "Contact Name": "Jane Smith",
            "Mobile Number": "987-654-3210",
            "Source URL": "https://example.com/2"
        }
    ]
    
    # Configure basic logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Test Excel export
    export_file = export_to_excel(sample_data)
    print(f"Exported to: {export_file}")